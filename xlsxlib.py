from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class Excel():
    
    # Library functions
    def create_workbook(self, name):
        
        name = name + '.xlsx'
        workbook = Workbook()
        workbook.save(name)  
        return
    
    def open_workbook(self, path):
        
        self.path = path + '.xlsx'        
        self.workbook = load_workbook(filename = self.path)
        self.worksheet = self.workbook.active
        return
	
    def save_workbook(self):

        self.workbook.save('sendlist.xlsx')   
        return

    def get_row_number(self):
        
        row_number = self.worksheet.max_row
        return row_number

    def read_cell(self, row_number, col_number):
            
        cell = self.worksheet.cell(row=row_number, column=col_number).value       
        return cell
	
    def write_cell(self, row_number, col_number, value):
        
        self.worksheet.cell(row=row_number, column=col_number).value = value      
        return
        
    def get_new_row(self, column):
        
        row_number = self.get_row_number()
        cell_data = []
        for i in range(1, row_number+1):

                data = self.read_cell(i, column)
                if data is not None:
                    cell_data.append(data)
            #cell_data = ' '.join(cell_data).split()
            
        new_empty_row = len(cell_data) + 1
        return new_empty_row

